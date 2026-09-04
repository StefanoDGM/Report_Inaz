import copy
import datetime as dt
import os
import subprocess
import re
import sys
import tempfile
import traceback
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:  # pragma: no cover
    tk = None
    filedialog = None
    messagebox = None
    ttk = None


SOURCE_SHEET_NAME = "Stampa Commesse Dipendente"
DETAIL_SHEET_NAME = "Stampa Commesse Dipendente"
SUMMARY_SHEET_NAME = "Riepilogo Viaggi"
ERRORS_SHEET_NAME = "Probabili errori"
DETAIL_OUTPUT_NAME = "ore_analitica.xls"
SUMMARY_OUTPUT_SUFFIX = "_riepilogo"
DAILY_SUMMARY_KEYWORD = "riepilogo giornaliero"
MAINTENANCE_TOLERANCE_MINUTES = 15.0
OTHER_EMPLOYEES_TOLERANCE_MINUTES = 15.0
MAINTENANCE_THEORETICAL_START_HOUR = 6.0

TOTAL_PROJ_MARKER = "Totale"
TRAVEL_KEYWORDS = ("COMMESSA", "CHIUSURA")
SUMMARY_SOURCE_COLUMNS = [
    (7, "Reparto"),
    (14, "Codice dipendente"),
    (15, "Nominativo"),
    (16, "Data"),
]


@dataclass
class GroupSummary:
    order: int
    first_row_idx: int
    base_row: list[Any]
    gross_worked_hours: float
    net_worked_hours: float
    office_hours: float
    travel_gross_hours: float
    travel_net_hours: float
    time_check_text: str = ""
    time_delta: Optional[float] = None

    @property
    def gross_ratio(self) -> float:
        if self.gross_worked_hours <= 0:
            return 0.0
        return self.travel_gross_hours / self.gross_worked_hours

    @property
    def net_ratio(self) -> float:
        if self.net_worked_hours <= 0:
            return 0.0
        return self.travel_net_hours / self.net_worked_hours

    @property
    def office_ratio(self) -> float:
        if self.gross_worked_hours <= 0:
            return 0.0
        return self.office_hours / self.gross_worked_hours

    @property
    def office_net_ratio(self) -> float:
        if self.net_worked_hours <= 0:
            return 0.0
        return self.office_hours / self.net_worked_hours


@dataclass
class ProcessingResult:
    detail_path: Path
    summary_path: Path


@dataclass
class DailySummaryEntry:
    gross_hours: float
    net_hours: float
    ordinary_hours: float
    overtime_hours: float
    has_recognized_work_hours: bool
    schedule_text: str
    row_text: str
    first_entry_time: Optional[float]

    @property
    def recognized_work_hours(self) -> float:
        return self.ordinary_hours + self.overtime_hours


def app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


def input_dir() -> Path:
    return app_root() / "input"


def output_dir() -> Path:
    return app_root() / "output"


def error_log_path() -> Path:
    return output_dir() / "ReportCommesse_error.log"


def shell_quote(value: Path | str) -> str:
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def convert_xlsx_to_xls(src_xlsx: Path, dst_xls: Path) -> None:
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$src = {shell_quote(src_xlsx)}
$dst = {shell_quote(dst_xls)}
$excel = New-Object -ComObject Excel.Application
$excel.DisplayAlerts = $false
$excel.Visible = $false
$workbook = $excel.Workbooks.Open($src)
$xlExcel5 = 39
$workbook.SaveAs($dst, $xlExcel5)
$workbook.Close($false)
$excel.Quit()
"""

    completed = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_script],
        capture_output=True,
        text=True,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "Impossibile convertire il file in formato .xls tramite Excel. "
            f"Dettagli: {completed.stderr.strip() or completed.stdout.strip() or 'errore sconosciuto'}"
        )


def ensure_workspace() -> None:
    input_dir().mkdir(parents=True, exist_ok=True)
    output_dir().mkdir(parents=True, exist_ok=True)


def write_error_log(exc: Exception) -> Path:
    ensure_workspace()
    log_path = error_log_path()
    timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_text = (
        f"[{timestamp}] {type(exc).__name__}: {exc}\n\n"
        f"{traceback.format_exc()}\n"
    )
    log_path.write_text(log_text, encoding="utf-8")
    return log_path


def is_daily_summary_file(path: Path) -> bool:
    return DAILY_SUMMARY_KEYWORD in normalize_text(path.stem)


def latest_input_file() -> Optional[Path]:
    files = sorted(
        [
            p
            for p in input_dir().glob("*.xlsx")
            if not p.name.startswith("~$")
            and not is_daily_summary_file(p)
        ],
        key=lambda p: (p.stat().st_mtime, p.name.lower()),
    )
    return files[-1] if files else None


def latest_daily_summary_file() -> Optional[Path]:
    files = sorted(
        [
            p
            for p in input_dir().glob("*.xlsx")
            if not p.name.startswith("~$")
            and is_daily_summary_file(p)
        ],
        key=lambda p: (p.stat().st_mtime, p.name.lower()),
    )
    return files[-1] if files else None


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def parse_duration(value: Any) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0

    if isinstance(value, dt.datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0

    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0

    if is_number(value):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    match = re.fullmatch(r"(?:(\d+):)?(\d{1,3}):(\d{2})", text)
    if match:
        hours_part = int(match.group(1) or match.group(2))
        minutes_part = int(match.group(3))
        return hours_part + minutes_part / 60.0

    match = re.fullmatch(r"(\d+(?:[.,]\d+)?)", text)
    if match:
        return float(match.group(1).replace(",", "."))

    return None


def parse_time_of_day(value: Any) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0

    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0

    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0

    text = str(value).strip()
    if not text:
        return None
    # Inaz sometimes exports time values with non-standard separators like
    # '#', '*', '.', or spaces. Normalize them before parsing.
    text = re.sub(r"[^\d]+", ":", text).strip(":")

    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        return hours + minutes / 60.0 + seconds / 3600.0

    return parse_duration(text)


def to_centesimal_hours(value: Any) -> Any:
    parsed = parse_duration(value)
    if parsed is None:
        return value
    return round(parsed, 5)


def is_date_value(value: Any) -> bool:
    return isinstance(value, (dt.datetime, dt.date))


def is_total_row(row: list[Any]) -> bool:
    for idx in (8, 9, 10, 11):
        if idx < len(row):
            cell = row[idx]
            if isinstance(cell, str) and cell.strip().lower() == TOTAL_PROJ_MARKER.lower():
                return True
    return False


def is_blank_row(row: list[Any]) -> bool:
    return all(str(cell or "").strip() == "" for cell in row)


def travel_row(row: list[Any]) -> bool:
    for idx in (8, 9, 10, 11):
        if idx < len(row):
            cell = row[idx]
            if isinstance(cell, str) and any(keyword in cell.upper() for keyword in TRAVEL_KEYWORDS):
                return True
    return False


def is_chiusura_row(row: list[Any]) -> bool:
    for idx in (8, 9, 10, 11):
        if idx < len(row):
            cell = row[idx]
            if isinstance(cell, str) and "CHIUSURA" in cell.upper():
                return True
    return False


def is_generic_commessa_row(row: list[Any]) -> bool:
    if len(row) <= 9:
        return False
    return normalize_text(row[9]) == "commessa"


def is_presence_row(row: list[Any]) -> bool:
    project = normalize_text(row[9]) if len(row) > 9 else ""
    cod_project = normalize_text(row[8]) if len(row) > 8 else ""
    cod_argomento = normalize_text(row[10]) if len(row) > 10 else ""
    argomento = normalize_text(row[11]) if len(row) > 11 else ""
    return not any((project, cod_project, cod_argomento, argomento))


def row_project_name(row: list[Any]) -> str:
    if len(row) <= 9:
        return ""
    return normalize_text(row[9])


def row_argument_name(row: list[Any]) -> str:
    if len(row) <= 11:
        return ""
    return normalize_text(row[11])


def row_argument_code(row: list[Any]) -> str:
    if len(row) <= 10:
        return ""
    return normalize_text(row[10])


def is_office_row(row: list[Any]) -> bool:
    if len(row) <= 9:
        return False
    project_name = str(row[9] or "").strip().casefold()
    return "sede ufficio" in project_name


def normalize_text(value: Any) -> str:
    return str(value or "").strip().casefold()


def is_generic_commessa_name(project_name: str) -> bool:
    return normalize_text(project_name) == "commessa"


def is_formazione_project(project_name: str) -> bool:
    return normalize_text(project_name) == "costi per formazione aziendale"


def is_valid_office_argument(argument_code: str, argument_name: str) -> bool:
    normalized_code = normalize_text(argument_code)
    normalized_name = normalize_text(argument_name)
    return (
        normalized_code in {"zzsede", "chiusura"}
        or normalized_name in {"attivita tecn in sede", "chiusura"}
    )


def is_valid_formazione_argument(argument_code: str, argument_name: str) -> bool:
    normalized_code = normalize_text(argument_code)
    normalized_name = normalize_text(argument_name)
    return (
        normalized_code in {"zzcosto", "chiusura"}
        or normalized_name in {"costo", "chiusura"}
    )


def row_mentions_ferie_or_rol(row: list[Any]) -> bool:
    text = " ".join(str(value or "") for value in row).casefold()
    return "ferie" in text or re.search(r"\br\.?o\.?l\.?\b", text) is not None


def daily_summary_mentions_ferie_or_rol(entry: DailySummaryEntry) -> bool:
    schedule_text = entry.schedule_text.casefold()
    row_text = entry.row_text
    return (
        "ferie" in row_text
        or re.search(r"\br\.?o\.?l\.?\b", row_text) is not None
        or "smart working" in row_text
        or "smartworking" in row_text
        or "smrwrk" in row_text
        or "smrwrk24" in row_text
        or "maternita" in row_text
        or "maternità" in row_text
        or "malattia" in row_text
        or re.search(r"\bmal\b", row_text) is not None
        or re.search(r"\bma1\b", row_text) is not None
        or re.search(r"\bma2\b", row_text) is not None
        or "rol" in schedule_text
        or "ferie" in schedule_text
        or "smart working" in schedule_text
        or "smartworking" in schedule_text
        or "smrwrk" in schedule_text
        or "smrwrk24" in schedule_text
        or "maternita" in schedule_text
        or "maternità" in schedule_text
        or "malattia" in schedule_text
        or re.search(r"\bmal\b", schedule_text) is not None
        or re.search(r"\bma1\b", schedule_text) is not None
        or re.search(r"\bma2\b", schedule_text) is not None
    )


def date_key(value: Any) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def is_workday(value: Any) -> bool:
    day = date_key(value)
    if day is None:
        return True
    return day.weekday() < 5


def find_daily_summary_header(ws, expected_labels: tuple[str, ...]) -> tuple[Optional[int], Optional[int]]:
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = normalize_text(ws.cell(row_idx, col_idx).value)
            if value in expected_labels:
                return row_idx, col_idx
    return None, None


def parse_daily_cause(value: Any) -> Optional[tuple[str, float]]:
    text = str(value or "").strip()
    if not text:
        return None

    match = re.match(r"^\s*(\S+)\s+(-?\d+(?:[.,]\d+)?)", text)
    if not match:
        return None

    code = re.sub(r"[^a-z0-9]", "", match.group(1).casefold())
    hours = float(match.group(2).replace(",", "."))
    return code, hours


def load_daily_summary_hours() -> dict[tuple[str, dt.date], DailySummaryEntry]:
    daily_path = latest_daily_summary_file()
    if daily_path is None:
        return {}

    wb = openpyxl.load_workbook(daily_path, data_only=True)
    ws = wb.active

    employee_header_row, employee_col = find_daily_summary_header(
        ws,
        ("nominativo", "dipendente", "dipendente nominativo", "nome dipendente", "cognome nome"),
    )
    date_header_row, date_col = find_daily_summary_header(
        ws,
        ("data", "giorno"),
    )
    schedule_header_row, schedule_col = find_daily_summary_header(
        ws,
        ("orario",),
    )

    if employee_col is None or date_col is None:
        return {}

    header_row = max(employee_header_row or 1, date_header_row or 1, schedule_header_row or 1)
    cause_cols = [
        col_idx
        for col_idx in range(1, ws.max_column + 1)
        if normalize_text(ws.cell(header_row, col_idx).value).startswith("cau")
    ]
    time_cols: list[tuple[int, int]] = []
    col_idx = 1
    while col_idx <= ws.max_column:
        if normalize_text(ws.cell(header_row, col_idx).value) == "e":
            next_value = normalize_text(ws.cell(header_row, col_idx + 1).value) if col_idx + 1 <= ws.max_column else ""
            if next_value == "u":
                time_cols.append((col_idx, col_idx + 1))
                col_idx += 2
                continue
        col_idx += 1

    daily_hours: dict[tuple[str, dt.date], DailySummaryEntry] = {}
    first_data_row = header_row + 1

    for row_idx in range(first_data_row, ws.max_row + 1):
        employee_name = str(ws.cell(row_idx, employee_col).value or "").strip()
        day_value = ws.cell(row_idx, date_col).value
        day_key = date_key(day_value)
        if not employee_name or day_key is None:
            continue

        worked_hours = 0.0
        first_entry_time: Optional[float] = None
        last_exit_time: Optional[float] = None
        for entry_col, exit_col in time_cols:
            entry_time = parse_time_of_day(ws.cell(row_idx, entry_col).value)
            exit_time = parse_time_of_day(ws.cell(row_idx, exit_col).value)
            if entry_time is None or exit_time is None:
                continue

            segment_hours = exit_time - entry_time
            if segment_hours < 0:
                segment_hours += 24.0
            worked_hours += segment_hours
            if first_entry_time is None:
                first_entry_time = entry_time
            last_exit_time = exit_time

        row_text = " ".join(str(ws.cell(row_idx, c).value or "") for c in range(1, ws.max_column + 1)).casefold()
        ordinary_hours = 0.0
        overtime_hours = 0.0
        has_recognized_work_hours = False
        for cause_col in cause_cols:
            parsed_cause = parse_daily_cause(ws.cell(row_idx, cause_col).value)
            if parsed_cause is None:
                continue
            cause_code, cause_hours = parsed_cause
            if cause_code == "ord":
                ordinary_hours += cause_hours
                has_recognized_work_hours = True
            elif cause_code.startswith("str"):
                overtime_hours += cause_hours
                has_recognized_work_hours = True

        gross_hours = worked_hours
        net_hours = worked_hours
        if first_entry_time is not None and last_exit_time is not None:
            gross_hours = last_exit_time - first_entry_time
            if gross_hours < 0:
                gross_hours += 24.0

        if is_workday(day_key):
            is_manutentori = "manutentori" in row_text
            if is_manutentori:
                net_hours = worked_hours
            else:
                actual_break = max(round(gross_hours - worked_hours, 5), 0.0)
                extra_break_needed = max(1.0 - actual_break, 0.0)
                net_hours = max(worked_hours - extra_break_needed, 0.0)
        else:
            net_hours = worked_hours

        key = (normalize_text(employee_name), day_key)
        schedule_text = str(ws.cell(row_idx, schedule_col).value or "").strip() if schedule_col else ""
        daily_hours[key] = DailySummaryEntry(
            gross_hours=round(gross_hours, 5),
            net_hours=round(net_hours, 5),
            ordinary_hours=round(ordinary_hours, 5),
            overtime_hours=round(overtime_hours, 5),
            has_recognized_work_hours=has_recognized_work_hours,
            schedule_text=schedule_text,
            row_text=row_text,
            first_entry_time=first_entry_time,
        )

    return daily_hours


def copy_cell_style(src, dst) -> None:
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy.copy(src.protection)


def clone_sheet_layout(src_ws, dst_ws) -> None:
    dst_ws.sheet_format = copy.copy(src_ws.sheet_format)
    dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy.copy(src_ws.page_margins)
    dst_ws.page_setup = copy.copy(src_ws.page_setup)
    dst_ws.print_options = copy.copy(src_ws.print_options)
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden
        dst_dim.bestFit = dim.bestFit
        dst_dim.outlineLevel = dim.outlineLevel
        dst_dim.collapsed = dim.collapsed
        dst_dim.min = dim.min
        dst_dim.max = dim.max


def row_quantity(value: Any) -> float:
    parsed = parse_duration(value)
    return round(parsed or 0.0, 5)


def distribute_amount(rows: list[dict[str, Any]], amount: float) -> float:
    if amount <= 0 or not rows:
        return amount

    total = sum(max(row["quantity"], 0.0) for row in rows)
    if total <= 0:
        return amount

    deducted = min(amount, total)
    for row in rows:
        share = deducted * (max(row["quantity"], 0.0) / total)
        row["quantity"] = round(row["quantity"] - share, 5)

    return round(amount - deducted, 5)


def adjust_lunch_break_for_group(rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    group_day = rows[0]["values"][16]
    if not is_workday(group_day):
        return

    manutentori_rows = [row for row in rows if row["is_manutentori"] and row["quantity"] > 0]
    if not manutentori_rows:
        return

    remaining = 1.0
    priority_buckets = [
        [row for row in manutentori_rows if row["bucket"] == "generic_commessa"],
        [row for row in manutentori_rows if row["bucket"] == "chiusura"],
        [row for row in manutentori_rows if row["bucket"] == "other_commessa"],
    ]

    for bucket_rows in priority_buckets:
        if remaining <= 0:
            break
        remaining = distribute_amount(bucket_rows, remaining)

    if remaining > 0:
        remaining = distribute_amount(manutentori_rows, remaining)

    for row in manutentori_rows:
        row["quantity"] = round(row["quantity"], 5)


def recognized_hours_with_adjustments(
    daily_entry: DailySummaryEntry,
    net_worked_hours: float,
    is_manutentori: bool,
) -> tuple[float, float, float]:
    recognized_hours = daily_entry.recognized_work_hours
    early_entry_credit = 0.0
    no_overtime_schedule_credit = 0.0

    if (
        normalize_text(daily_entry.schedule_text).endswith("n")
        and net_worked_hours > recognized_hours
    ):
        no_overtime_schedule_credit = net_worked_hours - recognized_hours
        return (
            round(net_worked_hours, 5),
            0.0,
            round(no_overtime_schedule_credit, 5),
        )

    if (
        is_manutentori
        and daily_entry.first_entry_time is not None
        and daily_entry.first_entry_time < MAINTENANCE_THEORETICAL_START_HOUR
        and (
            net_worked_hours - recognized_hours
            > MAINTENANCE_TOLERANCE_MINUTES / 60.0
        )
    ):
        early_entry_credit = (
            MAINTENANCE_THEORETICAL_START_HOUR - daily_entry.first_entry_time
        )

    return (
        round(recognized_hours + early_entry_credit, 5),
        round(early_entry_credit, 5),
        0.0,
    )


def collect_group_summaries(src_ws) -> OrderedDict[tuple[Any, Any, Any], dict[str, Any]]:
    grouped: OrderedDict[tuple[Any, Any, Any], dict[str, Any]] = OrderedDict()

    for row_idx in range(2, src_ws.max_row + 1):
        values = [src_ws.cell(row_idx, c).value for c in range(1, src_ws.max_column + 1)]
        if is_blank_row(values) or is_total_row(values):
            continue
        if len(values) < 18:
            continue

        employee_code = values[14]
        employee_name = values[15]
        day_value = values[16]

        if not employee_name or not is_date_value(day_value):
            continue

        key = (employee_code, employee_name, day_value)
        info = grouped.get(key)
        if info is None:
            grouped[key] = {
                "order": len(grouped),
                "first_row_idx": row_idx,
                "rows": [],
                "total_hours": 0.0,
                "office_hours": 0.0,
                "travel_gross_hours": 0.0,
            }
            info = grouped[key]

        quantity = parse_duration(values[17]) or 0.0
        row_info = {
            "row_idx": row_idx,
            "values": values,
            "quantity": quantity,
            "is_manutentori": str(values[7] or "").strip().upper() == "MANUTENTORI",
            "bucket": "other_commessa",
        }

        if is_generic_commessa_row(values):
            row_info["bucket"] = "generic_commessa"
        elif is_chiusura_row(values):
            row_info["bucket"] = "chiusura"

        info["rows"].append(row_info)
        info["total_hours"] += quantity
        if row_info["is_manutentori"] and is_office_row(values) and not is_chiusura_row(values):
            info["office_hours"] += quantity
        if row_info["is_manutentori"] and travel_row(values):
            info["travel_gross_hours"] += quantity

    return grouped


def build_detail_sheet(src_ws, dst_ws) -> dict[int, int]:
    clone_sheet_layout(src_ws, dst_ws)

    grouped = collect_group_summaries(src_ws)
    source_to_output_row: dict[int, int] = {}

    for c in range(1, src_ws.max_column + 1):
        src_cell = src_ws.cell(1, c)
        dst_cell = dst_ws.cell(1, c)
        dst_cell.value = src_cell.value
        copy_cell_style(src_cell, dst_cell)

    output_row = 2
    for info in grouped.values():
        detail_rows = [dict(row) for row in info["rows"]]
        adjust_lunch_break_for_group(detail_rows)

        for row in detail_rows:
            source_row_idx = row["row_idx"]
            source_to_output_row[source_row_idx] = output_row
            for c in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(source_row_idx, c)
                dst_cell = dst_ws.cell(output_row, c)
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)

                if c == 18:
                    dst_cell.value = round(row["quantity"], 5)
                    dst_cell.number_format = "0.00000"

            output_row += 1

    for c in range(1, src_ws.max_column + 1):
        dst_ws.column_dimensions[get_column_letter(c)].width = src_ws.column_dimensions[get_column_letter(c)].width

    last_col = get_column_letter(src_ws.max_column)
    dst_ws.auto_filter.ref = f"A1:{last_col}{max(output_row - 1, 1)}"
    return source_to_output_row


def build_summary_rows(
    src_ws,
    daily_summary_hours: dict[tuple[str, dt.date], DailySummaryEntry],
) -> list[GroupSummary]:
    grouped = collect_group_summaries(src_ws)

    summaries: list[GroupSummary] = []
    for info in grouped.values():
        base_row = list(info["rows"][0]["values"])
        net_rows = [dict(row) for row in info["rows"]]
        adjust_lunch_break_for_group(net_rows)
        net_worked_hours = round(sum(row["quantity"] for row in net_rows), 5)
        gross_travel = round(info["travel_gross_hours"], 5)
        is_manutentori = str(base_row[7] or "").strip().upper() == "MANUTENTORI"
        travel_net_hours = (
            round(
                sum(row["quantity"] for row in net_rows if travel_row(row["values"])),
                5,
            )
            if is_manutentori
            else 0.0
        )

        day_key = date_key(base_row[16])
        daily_entry = None
        time_check_text = ""
        time_delta = None
        if day_key is not None:
            summary_key = (normalize_text(base_row[15]), day_key)
            daily_entry = daily_summary_hours.get(summary_key)
        if daily_entry is not None:
            if daily_entry.has_recognized_work_hours:
                compared_hours, early_entry_credit, no_overtime_schedule_credit = recognized_hours_with_adjustments(
                    daily_entry,
                    net_worked_hours,
                    is_manutentori,
                )
                comparison_label = "ORD + STR netti"
                if no_overtime_schedule_credit > 0:
                    comparison_label += (
                        f" + straordinari esclusi da orario N "
                        f"{no_overtime_schedule_credit:.5f}"
                    )
                elif early_entry_credit > 0:
                    comparison_label += f" + anticipo prima delle 06:00 {early_entry_credit:.5f}"
            else:
                compared_hours = None
                comparison_label = ""

            if compared_hours is not None:
                time_delta = round(compared_hours - net_worked_hours, 5)
                time_check_text = (
                    f"{comparison_label} {compared_hours:.5f} - "
                    f"rendicontate nette {net_worked_hours:.5f}"
                )

        summaries.append(
            GroupSummary(
                order=info["order"],
                first_row_idx=info["first_row_idx"],
                base_row=base_row,
                gross_worked_hours=round(info["total_hours"], 5),
                net_worked_hours=net_worked_hours,
                office_hours=round(info["office_hours"], 5) if is_manutentori else 0.0,
                travel_gross_hours=gross_travel if is_manutentori else 0.0,
                travel_net_hours=travel_net_hours,
                time_check_text=time_check_text,
                time_delta=time_delta,
            )
        )

    return summaries


def build_summary_sheet(
    src_ws,
    dst_ws,
    daily_summary_hours: dict[tuple[str, dt.date], DailySummaryEntry],
) -> None:
    clone_sheet_layout(src_ws, dst_ws)

    headers = [label for _, label in SUMMARY_SOURCE_COLUMNS]
    headers.extend([
        "Ore lorde lavorate",
        "Ore nette lavorate",
        "Ore sede ufficio",
        "Ore viaggio lorde",
        "Ore viaggio nette",
        "% sede ufficio lorde",
        "% sede ufficio nette",
        "% viaggio lorde",
        "% viaggio nette",
        "Controllo ore INAZ",
        "Delta ore INAZ",
    ])

    summaries = build_summary_rows(src_ws, daily_summary_hours)

    title = dst_ws["A1"]
    title.value = "Riepilogo Viaggi"
    title.font = Font(bold=True, size=14)

    for c, header in enumerate(headers, start=1):
        cell = dst_ws.cell(4, c)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="999999")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, summary in enumerate(summaries, start=5):
        values = [summary.base_row[idx] for idx, _ in SUMMARY_SOURCE_COLUMNS]
        values.extend(
            [
                summary.gross_worked_hours,
                summary.net_worked_hours,
                summary.office_hours,
                summary.travel_gross_hours,
                summary.travel_net_hours,
                round(summary.office_ratio, 4),
                round(summary.office_net_ratio, 4),
                round(summary.gross_ratio, 4),
                round(summary.net_ratio, 4),
                summary.time_check_text,
                summary.time_delta,
            ]
        )

        for c, value in enumerate(values, start=1):
            cell = dst_ws.cell(row_idx, c)
            cell.value = value
            if c in (5, 6, 7, 8, 9):
                cell.number_format = "0.00000"
            elif c in (10, 11, 12, 13):
                cell.number_format = "0.0000%"
            elif c == 4:
                cell.number_format = "dd/mm/yyyy"
            elif c == 15 and value is not None:
                cell.number_format = "0.00000"

    dst_ws.freeze_panes = "A5"
    last_row = len(summaries) + 4
    last_col = get_column_letter(len(headers))
    dst_ws.auto_filter.ref = f"A4:{last_col}{last_row}"

    for c in range(1, len(headers) + 1):
        if c == 1:
            dst_ws.column_dimensions[get_column_letter(c)].width = 22
        elif c == 2:
            dst_ws.column_dimensions[get_column_letter(c)].width = 16
        elif c == 3:
            dst_ws.column_dimensions[get_column_letter(c)].width = 24
        elif c == 4:
            dst_ws.column_dimensions[get_column_letter(c)].width = 14
        elif c in (5, 6, 7, 8, 9):
            dst_ws.column_dimensions[get_column_letter(c)].width = 16
        elif c in (10, 11, 12, 13):
            dst_ws.column_dimensions[get_column_letter(c)].width = 14
        elif c == 14:
            dst_ws.column_dimensions[get_column_letter(c)].width = 40
        elif c == 15:
            dst_ws.column_dimensions[get_column_letter(c)].width = 16


def matching_detail_row(
    detail_ws,
    employee_name: Any,
    day_value: Any,
    preferred_row: Optional[int] = None,
) -> Optional[int]:
    expected_employee = normalize_text(employee_name)
    expected_day = date_key(day_value)

    def row_matches(row_idx: int) -> bool:
        if row_idx < 2 or row_idx > detail_ws.max_row:
            return False
        return (
            normalize_text(detail_ws.cell(row_idx, 16).value) == expected_employee
            and date_key(detail_ws.cell(row_idx, 17).value) == expected_day
        )

    if preferred_row is not None and row_matches(preferred_row):
        return preferred_row

    for row_idx in range(2, detail_ws.max_row + 1):
        if row_matches(row_idx):
            return row_idx
    return None


def build_errors_sheet(
    src_ws,
    detail_ws,
    dst_ws,
    detail_row_map: dict[int, int],
    daily_summary_hours: dict[tuple[str, dt.date], DailySummaryEntry],
) -> None:
    clone_sheet_layout(src_ws, dst_ws)

    grouped = collect_group_summaries(src_ws)
    error_rows: list[dict[str, Any]] = []

    for info in grouped.values():
        first_row_idx = info["first_row_idx"]
        first_row_values = info["rows"][0]["values"]
        day_label = first_row_values[16]
        employee_name = first_row_values[15]
        is_manutentori = str(first_row_values[7] or "").strip().upper() == "MANUTENTORI"
        total_hours = info["total_hours"]
        travel_gross_hours = info["travel_gross_hours"]

        # Check the original rows before applying the lunch-break deduction:
        # a COMMESSA row remains valid even when its adjusted quantity becomes zero.
        has_generic_commessa = any(
            is_generic_commessa_row(row["values"])
            for row in info["rows"]
        )
        if is_manutentori and not has_generic_commessa:
            mapped_row_idx = matching_detail_row(
                detail_ws,
                employee_name,
                day_label,
                detail_row_map.get(first_row_idx),
            )
            if mapped_row_idx is not None:
                error_rows.append(
                    {
                        "row_idx": mapped_row_idx,
                        "data": day_label,
                        "nominativo": employee_name,
                        "errore": "progetto COMMESSA mancante nella giornata",
                        "controllo": "",
                    }
                )

        net_rows = [dict(row) for row in info["rows"]]
        adjust_lunch_break_for_group(net_rows)
        net_worked_hours = round(sum(row["quantity"] for row in net_rows), 5)

        if total_hours > 0 and abs(travel_gross_hours - total_hours) <= 0.00001 and travel_gross_hours > 0:
            mapped_row_idx = matching_detail_row(
                detail_ws,
                employee_name,
                day_label,
                detail_row_map.get(first_row_idx),
            )
            if mapped_row_idx is not None:
                error_rows.append(
                    {
                        "row_idx": mapped_row_idx,
                        "data": day_label,
                        "nominativo": employee_name,
                        "errore": "ore viaggio 100%",
                        "controllo": "",
                    }
                )

        day_key = date_key(day_label)
        daily_entry = None
        if day_key is not None:
            daily_entry = daily_summary_hours.get((normalize_text(employee_name), day_key))

        if daily_entry is not None and not daily_summary_mentions_ferie_or_rol(daily_entry):
            if daily_entry.has_recognized_work_hours:
                compared_hours, early_entry_credit, no_overtime_schedule_credit = recognized_hours_with_adjustments(
                    daily_entry,
                    net_worked_hours,
                    is_manutentori,
                )
                tolerance_minutes = (
                    MAINTENANCE_TOLERANCE_MINUTES
                    if is_manutentori
                    else OTHER_EMPLOYEES_TOLERANCE_MINUTES
                )
                error_message = "ore nette rendicontate diverse da ORD + straordinari INAZ"
                comparison_label = "ORD + STR netti"
                if no_overtime_schedule_credit > 0:
                    comparison_label += (
                        f" + straordinari esclusi da orario N "
                        f"{no_overtime_schedule_credit:.5f}"
                    )
                elif early_entry_credit > 0:
                    comparison_label += f" + anticipo prima delle 06:00 {early_entry_credit:.5f}"
            else:
                compared_hours = None
                tolerance_minutes = 0.0
                error_message = ""
                comparison_label = ""

            if compared_hours is not None:
                delta_hours = round(compared_hours - net_worked_hours, 5)
                delta_minutes = abs(delta_hours) * 60.0
                if delta_minutes > tolerance_minutes + 0.00001:
                    mapped_row_idx = matching_detail_row(
                        detail_ws,
                        employee_name,
                        day_label,
                        detail_row_map.get(first_row_idx),
                    )
                    if mapped_row_idx is not None:
                        error_rows.append(
                            {
                                "row_idx": mapped_row_idx,
                                "data": day_label,
                                "nominativo": employee_name,
                                "errore": error_message,
                                "controllo": (
                                    f"delta {delta_hours:.5f} ({delta_minutes:.2f} min) | "
                                    f"{comparison_label} {compared_hours:.5f} - "
                                    f"rendicontate nette {net_worked_hours:.5f}"
                                ),
                            }
                        )

    for detail_row_idx in range(2, detail_ws.max_row + 1):
        values = [detail_ws.cell(detail_row_idx, c).value for c in range(1, detail_ws.max_column + 1)]
        if is_blank_row(values) or is_total_row(values):
            continue
        project_name = row_project_name(values)
        argument_name = row_argument_name(values)
        argument_code = row_argument_code(values)
        is_manutentori = str(values[7] or "").strip().upper() == "MANUTENTORI"

        if not project_name and not argument_name:
            error_rows.append(
                {
                    "row_idx": detail_row_idx,
                    "data": values[16],
                    "nominativo": values[15],
                    "errore": f"Riga {detail_row_idx} progetto e argomento mancanti",
                    "controllo": "",
                }
            )

        if project_name and project_name != "commessa" and not argument_name:
            error_rows.append(
                {
                    "row_idx": detail_row_idx,
                    "data": values[16],
                    "nominativo": values[15],
                    "errore": f'manca argomento per "{values[9]}"',
                    "controllo": "",
                }
            )
        if project_name and is_formazione_project(project_name) and not is_valid_formazione_argument(argument_code, argument_name):
            error_rows.append(
                {
                    "row_idx": detail_row_idx,
                    "data": values[16],
                    "nominativo": values[15],
                    "errore": 'argomento sbagliato per corso formazione: atteso "ZZCOSTO" oppure "CHIUSURA"',
                    "controllo": "",
                }
            )
        if (
            project_name
            and is_manutentori
            and is_office_row(values)
            and not is_valid_office_argument(argument_code, argument_name)
        ):
            error_rows.append(
                {
                    "row_idx": detail_row_idx,
                    "data": values[16],
                    "nominativo": values[15],
                    "errore": 'argomento sbagliato per sede ufficio: atteso "ATTIVITA TECN IN SEDE" oppure "CHIUSURA"',
                    "controllo": "",
                }
            )

    error_rows.sort(key=lambda item: (item["row_idx"], item["errore"]))

    title = dst_ws["A1"]
    title.value = "Probabili errori"
    title.font = Font(bold=True, size=14)

    headers = ["Numero riga", "Data", "Nominativo", "Errore", "Controllo ore INAZ"]
    for c, header in enumerate(headers, start=1):
        cell = dst_ws.cell(4, c)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F8D7DA")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="999999")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, item in enumerate(error_rows, start=5):
        dst_ws.cell(row_idx, 1).value = item["row_idx"]
        dst_ws.cell(row_idx, 2).value = item["data"]
        dst_ws.cell(row_idx, 3).value = item["nominativo"]
        dst_ws.cell(row_idx, 4).value = item["errore"]
        dst_ws.cell(row_idx, 5).value = item["controllo"]
        dst_ws.cell(row_idx, 2).number_format = "dd/mm/yyyy"

    dst_ws.freeze_panes = "A5"
    dst_ws.auto_filter.ref = f"A4:E{max(len(error_rows) + 4, 4)}"
    dst_ws.column_dimensions["A"].width = 14
    dst_ws.column_dimensions["B"].width = 14
    dst_ws.column_dimensions["C"].width = 24
    dst_ws.column_dimensions["D"].width = 48
    dst_ws.column_dimensions["E"].width = 56


def process_file(source_path: Path) -> ProcessingResult:
    wb = openpyxl.load_workbook(source_path)
    if SOURCE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Foglio '{SOURCE_SHEET_NAME}' non trovato nel file di input.")

    src_ws = wb[SOURCE_SHEET_NAME]

    detail_wb = Workbook()
    detail_default = detail_wb.active
    detail_wb.remove(detail_default)
    detail_wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    detail_ws = detail_wb.create_sheet(DETAIL_SHEET_NAME)
    detail_row_map = build_detail_sheet(src_ws, detail_ws)

    summary_wb = Workbook()
    summary_default = summary_wb.active
    summary_wb.remove(summary_default)
    summary_wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    summary_ws = summary_wb.create_sheet(SUMMARY_SHEET_NAME)
    errors_ws = summary_wb.create_sheet(ERRORS_SHEET_NAME)
    daily_summary_hours = load_daily_summary_hours()
    build_summary_sheet(src_ws, summary_ws, daily_summary_hours)
    build_errors_sheet(src_ws, detail_ws, errors_ws, detail_row_map, daily_summary_hours)

    detail_path = output_dir() / DETAIL_OUTPUT_NAME
    summary_path = output_dir() / f"{source_path.stem}{SUMMARY_OUTPUT_SUFFIX}.xlsx"
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_xlsx = Path(tmpdir) / f"{source_path.stem}_ore_analitica_temp.xlsx"
        detail_wb.save(temp_xlsx)
        convert_xlsx_to_xls(temp_xlsx, detail_path)
    summary_wb.save(summary_path)
    return ProcessingResult(detail_path=detail_path, summary_path=summary_path)


def choose_source_file() -> Optional[Path]:
    latest = latest_input_file()
    if latest:
        return latest
    if filedialog is None:
        return None

    selected = filedialog.askopenfilename(
        title="Seleziona il file di stampa commesse",
        initialdir=str(input_dir()),
        filetypes=[("Excel", "*.xlsx")],
    )
    if not selected:
        return None
    return Path(selected)


def run_gui() -> None:
    ensure_workspace()

    root = tk.Tk()
    root.title("Report Commesse")
    root.geometry("560x260")
    root.minsize(560, 260)

    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass

    main = ttk.Frame(root, padding=18)
    main.pack(fill="both", expand=True)

    title = ttk.Label(main, text="Elaborazione file stampa commesse", font=("Segoe UI", 14, "bold"))
    title.pack(anchor="w")

    info_var = tk.StringVar()
    latest = latest_input_file()
    if latest:
        info_var.set(f"File trovato: {latest.name}")
    else:
        info_var.set("Nessun file .xlsx trovato nella cartella input.")

    ttk.Label(main, textvariable=info_var, wraplength=520).pack(anchor="w", pady=(10, 12))

    output_var = tk.StringVar(value=f"Output: {output_dir()}")
    ttk.Label(main, textvariable=output_var, wraplength=520).pack(anchor="w", pady=(0, 12))

    def on_process() -> None:
        try:
            source = choose_source_file()
            if source is None:
                return

            result = process_file(source)
            messagebox.showinfo(
                "Completato",
                (
                    "File elaborato con successo.\n\n"
                    f"Input: {source.name}\n"
                    f"Dettaglio: {result.detail_path.name}\n"
                    f"Riepilogo: {result.summary_path.name}"
                ),
            )
            output_var.set(f"Creati: {result.detail_path.name} e {result.summary_path.name}")
        except Exception as exc:
            messagebox.showerror("Errore", str(exc))

    button_row = ttk.Frame(main)
    button_row.pack(fill="x", pady=(8, 0))
    ttk.Button(button_row, text="Elabora file", command=on_process).pack(side="left")

    def open_input_folder() -> None:
        os.startfile(str(input_dir()))

    def open_output_folder() -> None:
        os.startfile(str(output_dir()))

    ttk.Button(button_row, text="Apri input", command=open_input_folder).pack(side="left", padx=8)
    ttk.Button(button_row, text="Apri output", command=open_output_folder).pack(side="left")

    hint = ttk.Label(
        main,
        text="Il programma legge il file .xlsx piu recente da input, crea due file distinti in output e lascia il filtro manuale a Excel.",
        wraplength=520,
    )
    hint.pack(anchor="w", pady=(18, 0))

    root.mainloop()


def run_cli(argv: list[str]) -> int:
    ensure_workspace()

    source = None

    args = list(argv)
    if "--input" in args:
        idx = args.index("--input")
        if idx + 1 >= len(args):
            print("Manca il percorso dopo --input")
            return 2
        source = Path(args[idx + 1])

    if source is None:
        source = latest_input_file()

    if source is None or not source.exists():
        print("Nessun file trovato in input.")
        return 1

    try:
        result = process_file(source)
        print(f"Creati: {result.detail_path} | {result.summary_path}")
        return 0
    except Exception as exc:
        log_path = write_error_log(exc)
        print(f"Errore: {exc}")
        print(f"Dettagli salvati in: {log_path}")
        if getattr(sys, "frozen", False):
            try:
                input("Premi Invio per chiudere...")
            except EOFError:
                pass
        return 1


def main() -> int:
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        return run_cli(sys.argv[2:])

    if tk is None:
        return run_cli(sys.argv[1:])

    run_gui()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
